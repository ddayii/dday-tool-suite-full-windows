from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List

try:
    from openpyxl import load_workbook
except Exception:  # handled by caller
    load_workbook = None

IO_TYPES = ["DI", "DO", "RI", "RO", "GI", "GO", "UI", "UO", "SI", "SO", "WI", "WO", "AI", "AO", "R", "F", "M"]

ALIASES = {
    "REGISTER": "R", "REG": "R", "INTEGER REGISTER": "R",
    "FLAG": "F", "MARKER": "M",
}

@dataclass
class CommentRow:
    io_type: str
    number: int
    comment: str
    enabled: bool = True
    notes: str = ""


def normalize_type(value: str) -> str:
    text = str(value or "").strip().upper()
    return ALIASES.get(text, text)


def empty_data() -> Dict[str, List[CommentRow]]:
    return {key: [] for key in IO_TYPES}



def new_template_data() -> Dict[str, List[CommentRow]]:
    """Create an empty template data set for a new robot."""
    return empty_data()



def enabled_from_value(value) -> bool:
    """Return True only when Enabled is explicitly YES.

    Template validation limits user entries to blank, YES, or NO. Blank and
    NO are intentionally skipped during CSV/KAREL generation.
    """
    text = str(value if value is not None else "").strip().upper()
    return text == "YES"


def load_file(path: str | Path) -> Dict[str, List[CommentRow]]:
    p = Path(path)
    if p.suffix.lower() == ".csv":
        return load_roboguide_csv(p)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        return load_template_xlsx(p)
    raise ValueError("Input must be .csv, .xlsx, or .xlsm")


def load_roboguide_csv(path: str | Path) -> Dict[str, List[CommentRow]]:
    data = empty_data()
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        reader = csv.reader(f)
        rows = list(reader)

    start = 0
    if rows and len(rows[0]) >= 3:
        first = [c.strip().upper() for c in rows[0][:3]]
        if first == ["TYPE", "NUMBER", "COMMENT"]:
            start = 1

    for line_no, row in enumerate(rows[start:], start=start + 1):
        if len(row) < 2:
            continue
        io_type = normalize_type(row[0])
        if io_type not in data:
            continue
        try:
            number = int(str(row[1]).strip())
        except ValueError:
            continue
        comment = row[2].strip() if len(row) > 2 and row[2] is not None else ""
        notes = row[3].strip() if len(row) > 3 and row[3] is not None else ""
        data[io_type].append(CommentRow(io_type, number, comment, True, notes))

    sort_data(data)
    return data


def load_template_xlsx(path: str | Path) -> Dict[str, List[CommentRow]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read XLSX files.")

    # Load formulas instead of cached values so generated auto-entry rows still
    # import correctly even if Excel did not save recalculated formula results.
    wb = load_workbook(path, data_only=False)
    data = empty_data()

    for sheet_name in wb.sheetnames:
        io_type = normalize_type(sheet_name)
        if io_type not in data:
            continue
        ws = wb[sheet_name]
        headers = {str(c.value or "").strip().upper(): i + 1 for i, c in enumerate(ws[1])}
        number_col = headers.get("NUMBER")
        comment_col = headers.get("COMMENT")
        enabled_col = headers.get("ENABLED")
        notes_col = headers.get("NOTES")
        if not number_col or not comment_col:
            continue

        used_numbers = set()

        for r in range(2, ws.max_row + 1):
            comment = str(ws.cell(r, comment_col).value or "").strip()
            if not comment:
                continue

            raw_num = ws.cell(r, number_col).value
            number = None

            if isinstance(raw_num, str) and raw_num.startswith("="):
                number = (max(used_numbers) + 1) if used_numbers else 1
            elif raw_num not in (None, ""):
                try:
                    number = int(raw_num)
                except Exception:
                    number = None

            if number is None:
                number = (max(used_numbers) + 1) if used_numbers else 1

            raw_enabled = ws.cell(r, enabled_col).value if enabled_col else "YES"
            if isinstance(raw_enabled, str) and raw_enabled.startswith("="):
                enabled = True
            else:
                enabled = enabled_from_value(raw_enabled) if enabled_col else True

            notes = str(ws.cell(r, notes_col).value or "").strip() if notes_col else ""

            data[io_type].append(CommentRow(io_type, number, comment, enabled, notes))
            used_numbers.add(number)

    sort_data(data)
    return data


def sort_data(data: Dict[str, List[CommentRow]]) -> None:
    for key in data:
        data[key].sort(key=lambda x: x.number)


def iter_enabled(data: Dict[str, List[CommentRow]]) -> Iterable[CommentRow]:
    for io_type in IO_TYPES:
        for row in data.get(io_type, []):
            if row.enabled and row.comment:
                yield row


def count_rows(data: Dict[str, List[CommentRow]], enabled_only: bool = False) -> int:
    if enabled_only:
        return sum(1 for _ in iter_enabled(data))
    return sum(len(v) for v in data.values())
