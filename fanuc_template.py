from __future__ import annotations

from pathlib import Path
from datetime import datetime
from fanuc_io_parser import IO_TYPES

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter, quote_sheetname
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception:
    Workbook = None

HEADERS = ["Number", "Comment", "Enabled", "Notes"]
AUTO_ENTRY_ROWS = 500



def create_template(data, output_path: str | Path) -> Path:
    if Workbook is None:
        raise RuntimeError("openpyxl is required to create XLSX files.")

    path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    build_readme(ws)
    validation_ws = build_validation_sheet(wb)

    for io_type in IO_TYPES:
        ws = wb.create_sheet(io_type)
        build_sheet(ws, data.get(io_type, []), validation_ws.title)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    wb.save(path)
    return path


def build_readme(ws):
    ws["A1"] = "DDay Controls FANUC I/O Tool"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Purpose"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Use this workbook to edit FANUC comments and generate RoboGuide CSV or KAREL output."
    ws["A6"] = "Sheets"
    ws["A6"].font = Font(bold=True)
    ws["A7"] = ", ".join(IO_TYPES)
    ws["A9"] = "Columns"
    ws["A9"].font = Font(bold=True)
    ws["A10"] = "Number = I/O/register number. Comment = robot comment text. Enabled = blank/YES/NO dropdown. Notes = engineering notes only."
    ws["A11"] = "Each I/O tab uses the same template layout. Existing rows are filled from the source file, followed by blank entry rows."
    ws["A12"] = "To add a new item, type the Comment in the next blank row. Number auto-fills to the next available number and Enabled auto-fills YES."
    ws["A13"] = "You can overwrite Number or select blank/YES/NO in Enabled if the automatic values are not what you want."
    ws["A15"] = "Generated"
    ws["B15"] = datetime.now().strftime("%d-%b-%Y %H:%M")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80


def build_validation_sheet(wb):
    """Create hidden lookup values used by data-validation dropdowns."""
    ws = wb.create_sheet("_Validation")
    ws["A1"] = ""
    ws["A2"] = "YES"
    ws["A3"] = "NO"
    ws.sheet_state = "hidden"
    return ws


def build_sheet(ws, rows, validation_sheet_name: str):
    header_fill = PatternFill("solid", fgColor="303746")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="9CA8B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")

    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    data_rows = list(rows)

    # Use the same workbook template in all cases:
    #   - Existing source rows are populated first.
    #   - A bank of blank entry rows follows.
    #   - When the user types a Comment in a blank row, Excel formulas fill:
    #       Number  = next available number from the rows above
    #       Enabled = YES
    #     The user can still overwrite Number or Enabled if needed.
    total_rows = max(2, len(data_rows) + AUTO_ENTRY_ROWS + 1)

    for r in range(2, total_rows + 1):
        data_index = r - 2
        if data_index < len(data_rows):
            row = data_rows[data_index]
            values = [row.number, row.comment, "YES" if row.enabled else "NO", row.notes]
        else:
            # Formula rows remain visibly blank until a comment is entered.
            values = [
                f'=IF($B{r}<>"",IFERROR(MAX($A$1:A{r-1})+1,1),"")',
                "",
                f'=IF($B{r}<>"","YES","")',
                "",
            ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(r, c, val)
            cell.border = border

    widths = [12, 45, 12, 50]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    try:
        formula = f"={quote_sheetname(validation_sheet_name)}!$A$1:$A$3"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Enabled must be blank, YES, or NO."
        dv.errorTitle = "Invalid Enabled Value"
        dv.prompt = "Choose blank, YES, or NO."
        dv.promptTitle = "Enabled"
        ws.add_data_validation(dv)
        dv.add("C2:C1048576")
    except Exception:
        pass

    try:
        table_name = f"Tbl_{ws.title.replace(' ', '_')}"
        table = Table(displayName=table_name, ref=f"A1:D{total_rows}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    except Exception:
        ws.auto_filter.ref = f"A1:D{total_rows}"

    ws.freeze_panes = "A2"
